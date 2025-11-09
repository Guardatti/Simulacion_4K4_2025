"""
Exportador de resultados a Excel
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportadorExcel:
    """Exporta el vector de estado y las integraciones a Excel"""

    def __init__(self, vector_estado, metricas):
        self.vector_estado = vector_estado
        self.metricas = metricas

    def exportar(self, nombre_archivo="simulacion_biblioteca.xlsx"):
        """Exporta todo a un archivo Excel"""
        wb = openpyxl.Workbook()

        # Eliminar hoja por defecto
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Crear hojas
        self._crear_hoja_vector_estado(wb)
        self._crear_hoja_integraciones(wb)
        self._crear_hoja_metricas(wb)

        # Guardar archivo
        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado: {nombre_archivo}")

        return nombre_archivo

    def _crear_hoja_vector_estado(self, wb):
        """Crea la hoja con el vector de estado"""
        ws = wb.create_sheet("Vector de Estado")

        # Definir estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Fila", "Reloj (min)", "Evento", "Próximo Evento", "Tiempo Próx.",
            "Personas Dentro", "En Cola", "Leyendo", "Estado Biblioteca",
            "E1 Estado", "E1 Atendiendo", "E2 Estado", "E2 Atendiendo",
            "Total Llegadas", "Total Salidas", "Libros Pedidos", "Libros Devueltos",
            "Consultas", "No Entraron", "Tiempo Acum. Perm.", "Tiempo Cerrada"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Datos
        for idx, fila in enumerate(self.vector_estado, 2):
            ws.cell(row=idx, column=1).value = fila.numero_fila
            ws.cell(row=idx, column=2).value = round(fila.reloj, 2)
            ws.cell(row=idx, column=3).value = fila.evento

            # Próximo evento
            if fila.proximos_eventos:
                prox = fila.proximos_eventos[0]
                ws.cell(row=idx, column=4).value = prox['tipo']
                ws.cell(row=idx, column=5).value = round(prox['tiempo'], 2)

            ws.cell(row=idx, column=6).value = fila.biblioteca['personas_dentro']
            ws.cell(row=idx, column=7).value = len(fila.biblioteca['cola_atencion'])
            ws.cell(row=idx, column=8).value = len(fila.biblioteca['personas_leyendo'])
            ws.cell(row=idx, column=9).value = "CERRADA" if fila.biblioteca['cerrada'] else "ABIERTA"

            # Empleados
            for i, emp in enumerate(fila.biblioteca['empleados'][:2]):
                col_base = 10 + (i * 2)
                ws.cell(row=idx, column=col_base).value = emp['estado']
                ws.cell(row=idx, column=col_base + 1).value = emp['persona_atendiendo'] or ""

            # Acumuladores
            ws.cell(row=idx, column=14).value = fila.acumuladores['total_personas_llegadas']
            ws.cell(row=idx, column=15).value = fila.acumuladores['total_personas_salidas']
            ws.cell(row=idx, column=16).value = fila.acumuladores['total_libros_pedidos']
            ws.cell(row=idx, column=17).value = fila.acumuladores['total_libros_devueltos']
            ws.cell(row=idx, column=18).value = fila.acumuladores['total_consultas']
            ws.cell(row=idx, column=19).value = fila.acumuladores['personas_no_entraron_cerrada']
            ws.cell(row=idx, column=20).value = round(fila.acumuladores['tiempo_acumulado_permanencia'], 2)
            ws.cell(row=idx, column=21).value = round(fila.acumuladores['tiempo_biblioteca_cerrada'], 2)

            # Aplicar bordes
            for col in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col).border = border

        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Congelar primera fila
        ws.freeze_panes = "A2"

    def _crear_hoja_integraciones(self, wb):
        """Crea la hoja con las integraciones de Euler"""
        ws = wb.create_sheet("Integraciones Euler")

        # Estilos
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Headers
        ws.cell(row=1, column=1).value = "Fila"
        ws.cell(row=1, column=2).value = "Reloj (min)"
        ws.cell(row=1, column=3).value = "Persona"
        ws.cell(row=1, column=4).value = "Páginas Leídas"

        for col in range(1, 5):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        # Datos
        row = 2
        for fila in self.vector_estado:
            if fila.integraciones:
                for persona_id, integrador in fila.integraciones.items():
                    ws.cell(row=row, column=1).value = fila.numero_fila
                    ws.cell(row=row, column=2).value = round(fila.reloj, 2)
                    ws.cell(row=row, column=3).value = persona_id
                    ws.cell(row=row, column=4).value = round(integrador.obtener_valor_actual(), 2)
                    row += 1

        # Ajustar anchos
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.freeze_panes = "A2"

    def _crear_hoja_metricas(self, wb):
        """Crea la hoja con las métricas finales"""
        ws = wb.create_sheet("Métricas Finales")

        # Estilos
        title_font = Font(bold=True, size=14, color="1F4E78")
        label_font = Font(bold=True, size=11)
        value_font = Font(size=11)

        # Título
        ws.cell(row=1, column=1).value = "MÉTRICAS FINALES DE LA SIMULACIÓN"
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells("A1:B1")

        # Métricas
        row = 3
        metricas_labels = [
            ("Promedio de permanencia (min)", "promedio_permanencia"),
            ("% Tiempo biblioteca cerrada", "porcentaje_tiempo_cerrada"),
            ("Personas que no entraron (cerrada)", "personas_no_entraron"),
            ("Total personas llegadas", "total_personas_llegadas"),
            ("Total personas salidas", "total_personas_salidas"),
            ("Tiempo total simulado (min)", "tiempo_total_simulado")
        ]

        for label, key in metricas_labels:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = label_font

            valor = self.metricas[key]
            if isinstance(valor, float):
                ws.cell(row=row, column=2).value = round(valor, 2)
            else:
                ws.cell(row=row, column=2).value = valor
            ws.cell(row=row, column=2).font = value_font

            row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def exportar_historial_integraciones_detallado(self, nombre_archivo="integraciones_detalladas.xlsx"):
        """Exporta el historial completo de cada integración de Euler"""
        wb = openpyxl.Workbook()

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Recolectar todas las integraciones únicas
        integraciones_completas = {}

        for fila in self.vector_estado:
            if fila.integraciones:
                for persona_id, integrador in fila.integraciones.items():
                    if persona_id not in integraciones_completas:
                        integraciones_completas[persona_id] = integrador

        # Crear una hoja por cada persona que leyó
        for persona_id, integrador in integraciones_completas.items():
            ws = wb.create_sheet(persona_id)

            # Headers
            ws.cell(row=1, column=1).value = "Tiempo (min)"
            ws.cell(row=1, column=2).value = "Páginas Leídas"
            ws.cell(row=1, column=3).value = "K"

            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col in range(1, 4):
                ws.cell(row=1, column=col).fill = header_fill
                ws.cell(row=1, column=col).font = header_font

            # Datos del historial
            historial = integrador.obtener_historial()
            for idx, (tiempo, paginas) in enumerate(historial, 2):
                ws.cell(row=idx, column=1).value = round(tiempo, 4)
                ws.cell(row=idx, column=2).value = round(paginas, 4)
                ws.cell(row=idx, column=3).value = integrador.K

            # Ajustar anchos
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 10

        if integraciones_completas:
            wb.save(nombre_archivo)
            print(f"Archivo de integraciones detalladas generado: {nombre_archivo}")
        else:
            print("No hay integraciones para exportar.")
